import openpyxl as xl
# from openpyxl.styles import Font

wb = xl.load_workbook('example.xlsx')

sn = wb.sheetnames

print(sn)

sheet1 = wb['Sheet1']
cellA1 = sheet1['A1']

print(sheet1)
print(cellA1)

print(cellA1.value)
print(cellA1.row)
print(cellA1.column)
print(cellA1.coordinate)

print(sheet1.cell(1,3).value)

print(sheet1.max_row)
print(sheet1.max_column)

for i in range(1, sheet1.max_row):
    print(sheet1.cell(i,2).value)

print(xl.utils.get_column_letter(900))
print(xl.utils.column_index_from_string('AHP'))

for cr in sheet1['A1':'C3']:
    print(cr)
    for cc in cr:
        print(cc.coordinate, cc.value)

for cr in sheet1.iter_rows(min_row=1):
    print(cr[0].value)
    print(cr[1].value)
    print(cr[2].value)
    print(cr[3].value)
    print(cr[4].value)


# wb = xl.Workbook()
# ws = wb.active
# ws.title = 'First Sheet'